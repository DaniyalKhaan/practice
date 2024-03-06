import openpyxl
def loadDataFromExcel():
    wb = openpyxl.load_workbook('saveMe.xlsx')
    sheet = wb.active

    # Iterate through rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Update the "Sent" column for the sent row
        sheet.cell(row=1, column=1, value="AAAAa")
    # Save the changes to the Excel file
    wb.save('saveMe.xlsx')


# Example usage:
# Send additional 10 emails (assuming 10 emails are already sent)
loadDataFromExcel()

